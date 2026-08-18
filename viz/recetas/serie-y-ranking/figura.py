"""
Receta «serie + ranking» — caso de referencia: PIB trimestral e IpAEC 2026.

IZQUIERDA  PIB: variación trimestral. Las tres lecturas a la vez — interanual
           (barras), acumulada del año corrido y últimos 12 meses (líneas).
DERECHA    PIB: variación por actividad económica del primer semestre de 2026,
           ordenada de mayor a menor, con la línea del total.

Los dos trimestres de 2026 son estimación a partir del IpAEC del BCB (Reporte de
Inflación y Política Monetaria) aplicado sobre los niveles del PIB del INE. Eso
lo marca el ÁREA SOMBREADA, y solo eso: las series se dibujan continuas.

Ejecutar:  python figura.py   →  output/pib_ipaec_2026.png
Publicar:  python publicar.py (después de este)
"""
import sys
from pathlib import Path

import numpy as np
import openpyxl
from matplotlib.lines import Line2D
from matplotlib.patches import Patch

PUB = Path(__file__).resolve().parent
VIZ = PUB.parents[1]                       # .../galeria-populi/viz
sys.path.insert(0, str(VIZ))
import populi_style as ps                  # noqa: E402

FORMATO = "informe_panorama"
XLSX = PUB / "data" / "populi_pib_ipaec_2026.xlsx"
SALIDA = PUB / "output" / "pib_ipaec_2026.png"

DESDE = "2022-T1"             # arranque de la serie
EST_FROM = "2026-T1"          # primer trimestre estimado con el IpAEC

# Las BARRAS van en un solo color —el rojo de marca, serie principal del Banco—
# en los dos paneles. Las dos líneas se separan entre sí con el par validado.
C_BARRA = ps.col("rojo")
C_12M = ps.col("tinta")
C_ACUM = ps.col("serie_teal")

# Las barras van SEMITRANSPARENTES: las dos lineas cruzan por encima y con el
# rojo solido se perdian dentro de las barras.
BARRA_ALPHA = 0.72

# Escala tipografica de la figura, por debajo de la global del motor: aqui hay
# dos paneles en un lienzo, no uno, y con SIZES tal cual el texto se desborda.
S_EJE = 20        # numeros de los ejes          (motor: 25)
S_DATO = 19       # cifras y nombres del ranking (motor: 23)
S_LEYENDA = 19    # leyenda                      (motor: 23)
S_ROTULO = 25     # rotulo de cada panel         (motor: 29 * 0,88)

# Nombre corto para el eje del ranking.
CORTO = {
    "Actividad extractiva": "Extractiva",
    "Agricultura, ganadería, silvicultura y pesca": "Agropecuaria",
    "Suministro de electricidad, agua y recolección de desechos sólidos": "Electricidad y agua",
    "Administración pública, salud y educación de no mercado": "Admin. pública",
    "Actividades comunales sociales (**)": "Comunales y sociales",
    "Actividades de alojamiento y servicio de comidas y bebidas": "Alojamiento y comidas",
    "Actividades financieras y de seguros, inmobiliarias (*)": "Financieras e inmobiliarias",
    "Industrias manufactureras": "Manufactura",
    "Comercio": "Comercio",
    "Transporte y comunicaciones": "Transporte y comunicaciones",
    "Construcción": "Construcción",
}


# --------------------------------------------------------------------------- #
# Datos
# --------------------------------------------------------------------------- #
def leer():
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    ws = wb["PIB trimestral"]
    romano = {"I": 1, "II": 2, "III": 3, "IV": 4}
    etiq, inter, acum, doce = [], [], [], []
    for anio, trim, _per, _niv, a, ac, m12 in ws.iter_rows(
        min_row=9, max_row=ws.max_row, values_only=True
    ):
        if not isinstance(anio, int) or trim not in romano:
            continue
        etiq.append("%d-T%d" % (anio, romano[trim]))
        inter.append(None if a is None else a * 100)
        acum.append(None if ac is None else ac * 100)
        doce.append(None if m12 is None else m12 * 100)

    i0 = etiq.index(DESDE)
    etiq, inter, acum, doce = etiq[i0:], inter[i0:], acum[i0:], doce[i0:]
    if any(v is None for v in inter + acum + doce):
        raise ValueError("hay huecos en las series desde " + DESDE)

    hoja = [s for s in wb.sheetnames if s.startswith("IpAEC")][0]
    ws2 = wb[hoja]
    act, total = [], None
    for nombre, valor, _p, _c in ws2.iter_rows(min_row=5, max_row=ws2.max_row,
                                               values_only=True):
        if not isinstance(nombre, str):
            continue
        if nombre.startswith("Total IpAEC publicado"):
            total = float(valor)
        elif nombre in CORTO:
            act.append((CORTO[nombre], float(valor)))
    if len(act) != len(CORTO) or total is None:
        raise ValueError("la hoja del IpAEC no trae las 11 actividades o el total")

    return etiq, inter, acum, doce, sorted(act, key=lambda r: r[1]), total


etiq, inter, acum, doce, actividades, total = leer()
i_est = etiq.index(EST_FROM)
n = len(etiq)
x = np.arange(n, dtype=float)
print("%d trimestres (%s a %s) | %d actividades | total IpAEC %.2f%%"
      % (n, etiq[0], etiq[-1], len(actividades), total))


# --------------------------------------------------------------------------- #
# Marco de marca con componer() (eje dummy)
# --------------------------------------------------------------------------- #
W, H, sc = ps._spec(FORMATO)

fig, axd = ps.nueva_figura(FORMATO)
axd.set_xticks([]); axd.set_yticks([])
ps.componer(
    fig, axd,
    titulo="Bolivia: Evolución del PIB",
    subtitulo="Variación en porcentaje",
    fuente="Fuente: Instituto Nacional de Estadística (INE), PIB trimestral base 2017; "
           "Banco Central de Bolivia (BCB), Reporte de Inflación y Política Monetaria "
           "(IpAEC). Elaboración: Centro de Estudios POPULI · Carlos Aranda.",
    nota="Los dos trimestres de 2026 (área sombreada) son estimación a partir del IpAEC del BCB.",
    formato=FORMATO,
)
pos = axd.get_position()
axd.set_visible(False)

# --------------------------------------------------------------------------- #
# Dos columnas. El conjunto respeta los márgenes del título (izq = M, der = W−M);
# cada panel inserta su bloque de etiquetas Y al borde izquierdo de SU columna.
# --------------------------------------------------------------------------- #
head_h = 96 * sc / H           # rótulo del panel + fila de leyenda
col_gap = 110 * sc / W
# El ranking se lleva más ancho: carga el canal de los nombres de actividad.
colW_izq = (pos.width - col_gap) * 0.46
colW_der = (pos.width - col_gap) * 0.54
panelH = pos.height - head_h

ix0 = pos.x0
dx0 = pos.x0 + colW_izq + col_gap
ax_i = fig.add_axes([ix0, pos.y0, colW_izq, panelH])
ax_d = fig.add_axes([dx0, pos.y0, colW_der, panelH])

# ── Panel izquierdo: las tres lecturas del PIB ─────────────────────────────── #
# El área sombreada es lo ÚNICO que marca la estimación: las series van continuas.
ax_i.axvspan(i_est - 0.5, n - 0.5, color=ps.COLORS["cafe"], alpha=0.085,
             linewidth=0, zorder=1)

ax_i.bar(x, inter, width=0.66, color=C_BARRA, alpha=BARRA_ALPHA,
         linewidth=0, zorder=2)
ax_i.plot(x, acum, color=C_ACUM, linewidth=1.6 * sc, zorder=4, solid_capstyle="round")
ax_i.plot(x, doce, color=C_12M, linewidth=1.6 * sc, zorder=5, solid_capstyle="round")

ps.aplicar_estilo_ejes(ax_i, cero=True)
ax_i.grid(axis="y", color=ps.COLORS["borde"], linewidth=1.0 * sc,
          linestyle=(0, (1.6, 2.6)), alpha=1.0, zorder=0)
ax_i.set_ylim(-9.5, 15)
ax_i.set_xlim(-0.75, n - 0.25)
ax_i.set_yticks(range(-8, 15, 4))
ax_i.yaxis.set_major_formatter(ps.formateador_es(0, "%"))
anios = [(i, e[:4]) for i, e in enumerate(etiq) if e.endswith("-T1")]
ax_i.set_xticks([i for i, _ in anios])
ax_i.set_xticklabels([a for _, a in anios])

ax_i.text((i_est - 0.5 + n - 0.5) / 2, 14.2, "estimado", ha="center", va="top",
          color=ps.COLORS["cafe"],
          fontproperties=ps.fp(ps.BODY, S_DATO * sc), zorder=6)

# ── Panel derecho: el ranking ──────────────────────────────────────────────── #
nombres = [a for a, _ in actividades]
valores = [v for _, v in actividades]
y = np.arange(len(valores), dtype=float)

ax_d.barh(y, valores, height=0.62, color=C_BARRA, alpha=BARRA_ALPHA,
          linewidth=0, zorder=2)
ax_d.axvline(0, color=ps.COLORS["tinta"], linewidth=1.3 * sc, zorder=3)
ax_d.axvline(total, color=ps.COLORS["cafe"], linewidth=1.1 * sc,
             linestyle=(0, (3, 2.4)), zorder=4)

ps.aplicar_estilo_ejes(ax_d, grid_y=False)
ax_d.grid(axis="x", color=ps.COLORS["borde"], linewidth=1.0 * sc,
          linestyle=(0, (1.6, 2.6)), zorder=0)
ax_d.set_axisbelow(True)
ax_d.set_yticks(y)
ax_d.set_yticklabels(nombres)
ax_d.set_ylim(-1.5, len(valores) - 0.35)
ax_d.set_xlim(-31, 13)
ax_d.set_xticks(range(-30, 11, 10))
ax_d.xaxis.set_major_formatter(ps.formateador_es(0, "%"))

fp_dato = ps.fp("JetBrains Mono SemiBold", S_DATO * sc)
for yi, v in zip(y, valores):
    dentro = abs(v) >= 12
    lado = 1 if v >= 0 else -1
    ax_d.annotate(("+" if v > 0 else "") + ps.es_num(v, 1) + "%", (v, yi),
                  xytext=((-lado if dentro else lado) * 8 * sc, 0),
                  textcoords="offset points",
                  ha=("right" if v >= 0 else "left") if dentro
                     else ("left" if v >= 0 else "right"),
                  va="center", zorder=6, fontproperties=fp_dato,
                  color=ps.contraste_texto(C_BARRA) if dentro
                        else ps.COLORS["cafe_oscuro"])

ax_d.annotate("Total " + ps.es_num(total, 1) + "%", (total, -1.15),
              xytext=(-9 * sc, 0), textcoords="offset points",
              ha="right", va="center", color=ps.COLORS["cafe"],
              fontproperties=ps.fp("Inter Bold", S_DATO * sc), zorder=6)

# --------------------------------------------------------------------------- #
# Encuadre: el bloque de etiquetas Y de cada panel se inserta al borde izquierdo
# de SU columna (el del panel izquierdo coincide con el margen del título).
# --------------------------------------------------------------------------- #
fp_eje = ps.fp(ps.MONO, S_EJE * sc)         # cifras de los ejes
fp_nom = ps.fp(ps.BODY, S_DATO * sc)        # nombres de actividad: fuente de texto
for lbl in ax_i.get_xticklabels() + ax_i.get_yticklabels() + ax_d.get_xticklabels():
    lbl.set_fontproperties(fp_eje)
for lbl in ax_d.get_yticklabels():
    lbl.set_fontproperties(fp_nom)
fig.canvas.draw()
_rend = fig.canvas.get_renderer()


def _inset_y(ax, col_left, col_right):
    """Inserta el bloque de etiquetas Y al borde izquierdo de su columna."""
    anchos = [l.get_window_extent(_rend).width for l in ax.get_yticklabels()
              if l.get_text()]
    lw = max(anchos) if anchos else 0.0
    spine = col_left * W + lw + 8 * sc * ps.DPI / 72.0
    p = ax.get_position()
    ax.set_position([spine / W, p.y0, col_right - spine / W, p.height])


_inset_y(ax_i, ix0, ix0 + colW_izq)
_inset_y(ax_d, dx0, dx0 + colW_der)

# El ranking rotula sus barras positivas FUERA del area de trazado, asi que la
# cifra se salia del encuadre. El margen derecho lo fija el wordmark (W - M), y
# es contra ESE limite que hay que encoger. Se mide el desborde real y se corrige;
# encoger mueve las barras, asi que converge en un par de pasadas.
_lim = (dx0 + colW_der) * W
for _ in range(4):
    fig.canvas.draw()
    _rend = fig.canvas.get_renderer()
    _over = max(t.get_window_extent(_rend).x1 for t in ax_d.texts) - _lim
    if _over <= 1:
        break
    _p = ax_d.get_position()
    ax_d.set_position([_p.x0, _p.y0, _p.width - _over / W, _p.height])

# ── Leyenda: fila propia sobre el panel izquierdo, fuera del área de trazado ── #
leg = ax_i.legend(
    handles=[Patch(facecolor=C_BARRA, alpha=BARRA_ALPHA, edgecolor="none",
                   label="Interanual"),
             Line2D([0], [0], color=C_ACUM, linewidth=1.6 * sc, label="Acumulada"),
             Line2D([0], [0], color=C_12M, linewidth=1.6 * sc, label="A 12 meses")],
    loc="lower left", bbox_to_anchor=(0, 1.01), ncol=3, frameon=False,
    handlelength=1.3, handletextpad=0.45, columnspacing=1.3, borderaxespad=0,
    prop=ps.fp(ps.BODY, S_LEYENDA * sc))
for t in leg.get_texts():
    t.set_color(ps.COLORS["cafe"])

# Rótulo de cada panel, al borde de su columna, por encima de la leyenda.
fp_rot = ps.fp("Inter Bold", S_ROTULO * sc)
y_rot = pos.y0 + panelH + 58 * sc / H
fig.text(ix0, y_rot, "PIB: Variación Trimestral",
         fontproperties=fp_rot, color=ps.COLORS["cafe_oscuro"], va="bottom", ha="left")
fig.text(dx0, y_rot, "PIB: Variación por Actividad Económica (1er Semestre)",
         fontproperties=fp_rot, color=ps.COLORS["cafe_oscuro"], va="bottom", ha="left")

ps.guardar(fig, str(SALIDA), formato=FORMATO, thumb=False)
print("OK render")
